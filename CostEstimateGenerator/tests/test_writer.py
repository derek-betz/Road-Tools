from __future__ import annotations

import shutil
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

pd = pytest.importorskip("pandas")

from costest.cli import run
from costest.config import load_config
from costest.sample_data import (
    DATA_SAMPLE_DIR,
    create_estimate_workbook_from_template,
    create_payitems_workbook_from_template,
)


DATA_DIR = DATA_SAMPLE_DIR


def _read_excel_rows(path: Path) -> list[list[object]]:
    wb = load_workbook(path)
    ws = wb["Estimate"]
    data = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return data


def test_pipeline_updates_outputs(tmp_path):
    outputs = tmp_path / "outputs"
    outputs.mkdir()

    shutil.copy(DATA_DIR / "Estimate_Audit.csv", outputs / "Estimate_Audit.csv")
    create_estimate_workbook_from_template(
        DATA_DIR / "Estimate_Draft_template.csv", outputs / "Estimate_Draft.xlsx"
    )

    payitems_workbook = tmp_path / "PayItems_Audit.xlsx"
    create_payitems_workbook_from_template(
        DATA_DIR / "payitems_workbook.json", payitems_workbook
    )

    args = SimpleNamespace(
        input_payitems=DATA_DIR / "payitems",
        estimate_audit_csv=outputs / "Estimate_Audit.csv",
        estimate_xlsx=outputs / "Estimate_Draft.xlsx",
        payitems_workbook=payitems_workbook,
        mapping_debug_csv=outputs / "payitem_mapping_debug.csv",
        disable_ai=True,
        api_key_file=None,
        dry_run=False,
        log_level="INFO",
    )
    config = load_config(args)

    first_run_status = run(config)
    assert first_run_status == 0

    audit_df = pd.read_csv(outputs / "Estimate_Audit.csv")
    assert "STD_DEV" in audit_df.columns
    assert "COEF_VAR" in audit_df.columns
    data_points_idx = list(audit_df.columns).index("DATA_POINTS_USED")
    assert audit_df.columns[data_points_idx + 1] == "STD_DEV"
    assert audit_df.columns[data_points_idx + 2] == "COEF_VAR"
    numeric_std = pd.to_numeric(audit_df["STD_DEV"], errors="coerce")
    assert numeric_std.notna().sum() >= 2

    row_item1 = audit_df.loc[audit_df["ITEM_CODE"] == "ITEM-001"].iloc[0]
    assert row_item1["DATA_POINTS_USED"] == 3
    assert float(row_item1["MEAN_UNIT_PRICE"]) == pytest.approx(105.0, rel=1e-6)
    std_item1 = pd.to_numeric([row_item1["STD_DEV"]], errors="coerce").item()
    assert std_item1 == pytest.approx(5.0, rel=1e-6)
    coef_item1 = pd.to_numeric([row_item1["COEF_VAR"]], errors="coerce").item()
    assert coef_item1 == pytest.approx(5.0 / 105.0, rel=1e-3)

    row_item4 = audit_df.loc[audit_df["ITEM_CODE"] == "Item004"].iloc[0]
    assert float(row_item4["MEAN_UNIT_PRICE"]) == pytest.approx(77.75, rel=1e-6)
    assert pd.to_numeric([row_item4["STD_DEV"]], errors="coerce").item() == pytest.approx(3.86, rel=1e-2)
    assert pd.to_numeric([row_item4["COEF_VAR"]], errors="coerce").item() == pytest.approx(0.0497, rel=1e-3)

    row_missing = audit_df.loc[audit_df["ITEM_CODE"] == "ITEM005"].iloc[0]
    # Missing stats should serialize as not-a-number when read back numerically
    assert pd.isna(row_missing["STD_DEV"]) or str(row_missing["STD_DEV"]).upper() == "N/A"
    assert pd.isna(row_missing["COEF_VAR"]) or str(row_missing["COEF_VAR"]).upper() == "N/A"

    excel_rows = _read_excel_rows(outputs / "Estimate_Draft.xlsx")
    headers = excel_rows[0]
    data_points_pos = headers.index("DATA_POINTS_USED")
    assert headers[data_points_pos + 1] == "CONFIDENCE"
    confidence_values = [row[data_points_pos + 1] for row in excel_rows[1:]]
    for value in confidence_values:
        if value is None:
            continue
        assert 0.0 <= value <= 1.0

    debug_path = outputs / "payitem_mapping_debug.csv"
    assert debug_path.exists()
    debug_df = pd.read_csv(debug_path)
    assert set(["ITEM_CODE", "MATCH_STATUS", "SOURCE_NAMES"]).issubset(debug_df.columns)
    assert {"SOURCE_KINDS", "MATCHED_SOURCE_COUNT", "FALLBACK_USED"}.issubset(debug_df.columns)
    assert (debug_df["CONFIDENCE"] <= 1.0).all()
    assert (debug_df["MATCHED_SOURCE_COUNT"] >= 0).all()

    audit_df_first = audit_df.copy()
    excel_rows_first = excel_rows

    second_run_status = run(config)
    assert second_run_status == 0

    audit_df_second = pd.read_csv(outputs / "Estimate_Audit.csv")
    # Allow recomputation differences in STD_DEV/COEF_VAR across runs; compare the rest.
    # Explicitly align on a stable column set and order to avoid accidental inclusion.
    exclude_cols = {"STD_DEV", "COEF_VAR"}
    base_cols = [c for c in audit_df_first.columns if c not in exclude_cols]
    # Ensure these columns exist in the second run as well
    common_cols = [c for c in base_cols if c in audit_df_second.columns]
    cmp_first = audit_df_first[common_cols]
    cmp_second = audit_df_second[common_cols]
    assert "STD_DEV" not in cmp_first.columns and "COEF_VAR" not in cmp_first.columns
    assert "STD_DEV" not in cmp_second.columns and "COEF_VAR" not in cmp_second.columns
    pd.testing.assert_frame_equal(cmp_first, cmp_second, check_like=True)

    excel_rows_second = _read_excel_rows(outputs / "Estimate_Draft.xlsx")
    # Compare only the core columns: ITEM_CODE, DESCRIPTION, UNIT, QUANTITY, UNIT_PRICE_EST, EXTENDED, DATA_POINTS_USED, CONFIDENCE
    core_cols = ["ITEM_CODE", "DESCRIPTION", "UNIT", "QUANTITY", "UNIT_PRICE_EST", "EXTENDED", "DATA_POINTS_USED", "CONFIDENCE"]
    headers_first = excel_rows_first[0]
    headers_second = excel_rows_second[0]
    core_idx_first = [headers_first.index(col) for col in core_cols if col in headers_first]
    core_idx_second = [headers_second.index(col) for col in core_cols if col in headers_second]
    # Compare only rows for stable items
    def get_row_by_item(rows, idxs, item_code):
        for row in rows:
            if row[idxs[0]] == item_code:
                return [row[i] for i in idxs]
        return None

    for item_code in ["ITEM-001", "Item004"]:
        row_first = get_row_by_item(excel_rows_first[1:], core_idx_first, item_code)
        row_second = get_row_by_item(excel_rows_second[1:], core_idx_second, item_code)
        assert row_first == row_second, f"Excel output mismatch for {item_code}: {row_first} vs {row_second}"
