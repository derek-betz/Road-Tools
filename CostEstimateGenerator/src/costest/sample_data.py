"""Helpers for generating sample workbooks from text templates."""
from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

from openpyxl import Workbook

DATA_SAMPLE_DIR = Path(__file__).resolve().parents[2] / "data_sample"


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def load_template_rows(csv_path: Path) -> List[List[str]]:
    with open(csv_path, newline="", encoding="utf-8") as handle:
        reader = csv.reader(handle)
        return [row for row in reader]


def create_estimate_workbook_from_template(template_csv: Path, destination: Path) -> Path:
    rows = load_template_rows(template_csv)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Estimate"
    for row in rows:
        sheet.append(row)
    _ensure_parent(destination)
    workbook.save(destination)
    workbook.close()
    return destination


def _iter_workbook_templates(template_json: Path) -> Iterable[tuple[str, Dict[str, Sequence[object]]]]:
    with open(template_json, encoding="utf-8") as handle:
        data = json.load(handle)
    for sheet_name, payload in data.items():
        columns = payload.get("columns", [])
        rows = payload.get("rows", [])
        yield sheet_name, {"columns": columns, "rows": rows}


def create_payitems_workbook_from_template(template_json: Path, destination: Path) -> Path:
    workbook = Workbook()
    first = True
    for sheet_name, payload in _iter_workbook_templates(template_json):
        if first:
            sheet = workbook.active
            sheet.title = sheet_name
            first = False
        else:
            sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(list(payload["columns"]))
        for row in payload["rows"]:
            values = list(row)
            sheet.append([value if value is not None else "" for value in values])
    _ensure_parent(destination)
    workbook.save(destination)
    workbook.close()
    return destination


__all__ = [
    "DATA_SAMPLE_DIR",
    "create_estimate_workbook_from_template",
    "create_payitems_workbook_from_template",
    "load_template_rows",
]
