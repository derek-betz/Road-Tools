"""Writers that update Estimate outputs in-place."""
from __future__ import annotations

import logging
import math
import shutil
import tempfile
from pathlib import Path
from typing import Dict, Iterable, Mapping

import pandas as pd
from openpyxl import load_workbook

from .mapping import normalize_key
from .stats import StatsSummary

LOGGER = logging.getLogger(__name__)


def _find_item_code_column(df: pd.DataFrame) -> str:
    for column in df.columns:
        lowered = str(column).lower()
        if "item" in lowered and "code" in lowered:
            return column
    raise KeyError("Could not identify ITEM_CODE column in Estimate_Audit.csv")


def _ensure_columns(df: pd.DataFrame, after: str, new_columns: Iterable[str]) -> pd.DataFrame:
    columns = list(df.columns)
    for column in new_columns:
        if column not in columns:
            df[column] = ""
            columns.append(column)
    # Reorder columns
    for column in new_columns:
        if column in columns:
            columns.remove(column)
    if after not in columns:
        raise KeyError(f"Column {after} not found while updating Estimate_Audit.csv")
    insert_index = columns.index(after) + 1
    for offset, column in enumerate(new_columns):
        columns.insert(insert_index + offset, column)
    return df.loc[:, columns]


def update_estimate_audit(csv_path: Path, stats_by_item: Mapping[str, StatsSummary], dry_run: bool = False) -> None:
    df = pd.read_csv(csv_path)
    item_column = _find_item_code_column(df)
    df = _ensure_columns(df, "DATA_POINTS_USED", ["STD_DEV", "COEF_VAR"])

    for idx, row in df.iterrows():
        item_code = str(row[item_column])
        stats = stats_by_item.get(normalize_key(item_code))
        if not stats:
            stats = StatsSummary(0, 0.0, 0.0, math.inf, 0.0, notes="No history", fallback_used=True)
        df.at[idx, "DATA_POINTS_USED"] = int(stats.data_points)
        if stats.data_points:
            df.at[idx, "MEAN_UNIT_PRICE"] = round(stats.mean, 2)

        if stats.data_points == 0:
            df.at[idx, "STD_DEV"] = "N/A"
            df.at[idx, "COEF_VAR"] = "N/A"
        else:
            df.at[idx, "STD_DEV"] = round(stats.std_dev, 2)
            df.at[idx, "COEF_VAR"] = (
                round(stats.coef_var, 4) if math.isfinite(stats.coef_var) else "N/A"
            )

    if dry_run:
        LOGGER.info("Dry-run enabled; Estimate_Audit.csv not written.")
        return

    tmp = tempfile.NamedTemporaryFile("w", delete=False, suffix=".csv", newline="")
    try:
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(tmp.name, index=False)
        tmp.close()
        shutil.move(tmp.name, csv_path)
    finally:
        Path(tmp.name).unlink(missing_ok=True)

    LOGGER.info("Updated %s", csv_path)


def update_estimate_draft(xlsx_path: Path, stats_by_item: Mapping[str, StatsSummary], dry_run: bool = False) -> None:
    wb = load_workbook(xlsx_path)
    if "Estimate" not in wb.sheetnames:
        raise KeyError("Expected 'Estimate' sheet in Estimate_Draft workbook")
    ws = wb["Estimate"]

    headers = [cell.value if cell.value is not None else "" for cell in ws[1]]

    def find_header(target: str) -> int:
        for idx, header in enumerate(headers, start=1):
            if isinstance(header, str) and header.strip().lower() == target.lower():
                return idx
        raise KeyError(f"Column {target} not found in Estimate sheet")

    item_col_idx = find_header("ITEM_CODE")
    data_points_idx = find_header("DATA_POINTS_USED")
    mean_idx = find_header("MEAN_UNIT_PRICE")

    try:
        confidence_idx = find_header("CONFIDENCE")
    except KeyError:
        ws.insert_cols(data_points_idx + 1)
        ws.cell(row=1, column=data_points_idx + 1, value="CONFIDENCE")
        headers.insert(data_points_idx, "CONFIDENCE")
        confidence_idx = data_points_idx + 1
        # Indices after insertion shift for columns to the right
        if mean_idx >= confidence_idx:
            mean_idx += 1

    for row_idx in range(2, ws.max_row + 1):
        item_code = ws.cell(row=row_idx, column=item_col_idx).value
        if item_code is None:
            continue
        stats = stats_by_item.get(normalize_key(str(item_code)))
        if not stats:
            stats = StatsSummary(0, 0.0, 0.0, math.inf, 0.0, notes="No history", fallback_used=True)
        ws.cell(row=row_idx, column=data_points_idx, value=int(stats.data_points))
        if stats.data_points:
            ws.cell(row=row_idx, column=mean_idx, value=float(round(stats.mean, 2)))
        ws.cell(row=row_idx, column=confidence_idx, value=float(round(stats.confidence, 4)))

    if dry_run:
        LOGGER.info("Dry-run enabled; Estimate_Draft.xlsx not written.")
        wb.close()
        return

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    try:
        xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(tmp.name)
        wb.close()
        tmp.close()
        shutil.move(tmp.name, xlsx_path)
    finally:
        Path(tmp.name).unlink(missing_ok=True)

    LOGGER.info("Updated %s", xlsx_path)


def write_mapping_debug(csv_path: Path, records: Iterable[Dict[str, object]], dry_run: bool = False) -> None:
    if dry_run:
        LOGGER.info("Dry-run enabled; mapping debug file not written.")
        return

    df = pd.DataFrame(list(records))
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(csv_path, index=False)
    LOGGER.info("Wrote mapping debug report to %s", csv_path)


__all__ = [
    "update_estimate_audit",
    "update_estimate_draft",
    "write_mapping_debug",
]
